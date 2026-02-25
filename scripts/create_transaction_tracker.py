#!/usr/bin/env python3
"""
Creates a properly formatted Excel transaction tracker for Hansen Web Services
Run this script to generate HWS_Transaction_Tracker.xlsx
"""

import pandas as pd
from datetime import datetime
import os

def create_transaction_tracker():
    # Define the columns
    columns = [
        'Date',
        'Transaction ID',
        'Client Name',
        'Transaction Type',
        'Category',
        'Description',
        'Payment Method',
        'Amount',
        'Fees',
        'Net Amount',
        'Status',
        'Invoice/Receipt #',
        'Project',
        'Notes'
    ]

    # Create example data
    example_data = [
        {
            'Date': '01/20/2026',
            'Transaction ID': 'TXN-001',
            'Client Name': 'ABC Company',
            'Transaction Type': 'Income',
            'Category': 'Down Payment',
            'Description': 'Professional Package - 50% deposit',
            'Payment Method': 'Stripe',
            'Amount': 247.50,
            'Fees': 7.47,
            'Net Amount': 240.03,
            'Status': 'Completed',
            'Invoice/Receipt #': 'INV-2026-001',
            'Project': 'ABC Company Website',
            'Notes': 'Client signed contract 01/19'
        },
        {
            'Date': '01/20/2026',
            'Transaction ID': 'EXP-001',
            'Client Name': '',
            'Transaction Type': 'Expense',
            'Category': 'Software Subscription',
            'Description': 'GitHub Copilot monthly',
            'Payment Method': 'Credit Card',
            'Amount': 10.00,
            'Fees': 0.00,
            'Net Amount': 10.00,
            'Status': 'Completed',
            'Invoice/Receipt #': 'RCT-GH-456',
            'Project': '',
            'Notes': 'Business development tool'
        },
        {
            'Date': '01/22/2026',
            'Transaction ID': 'TXN-002',
            'Client Name': 'XYZ Shop',
            'Transaction Type': 'Income',
            'Category': 'Care Plan Subscription',
            'Description': 'Basic Care - Monthly',
            'Payment Method': 'Stripe',
            'Amount': 25.00,
            'Fees': 1.03,
            'Net Amount': 23.97,
            'Status': 'Completed',
            'Invoice/Receipt #': 'INV-2026-002',
            'Project': 'XYZ Shop Website',
            'Notes': 'First month of care plan'
        }
    ]

    # Create DataFrame with examples
    df_examples = pd.DataFrame(example_data)

    # Create empty DataFrame for user entries (20 blank rows)
    df_blank = pd.DataFrame(columns=columns, index=range(20))

    # Combine examples and blank rows
    df = pd.concat([df_examples, df_blank], ignore_index=True)

    # Get the directory where the script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    # Go up one level to HWS root
    project_root = os.path.dirname(script_dir)
    # Save in business-docs folder
    output_path = os.path.join(project_root, 'business-docs', 'HWS_Transaction_Tracker.xlsx')

    # Create Excel writer object
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # Write main transactions sheet
        df.to_excel(writer, sheet_name='Transactions', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Transactions']

        # Set column widths
        column_widths = {
            'A': 12,  # Date
            'B': 15,  # Transaction ID
            'C': 20,  # Client Name
            'D': 16,  # Transaction Type
            'E': 25,  # Category
            'F': 35,  # Description
            'G': 16,  # Payment Method
            'H': 12,  # Amount
            'I': 10,  # Fees
            'J': 12,  # Net Amount
            'K': 12,  # Status
            'L': 18,  # Invoice #
            'M': 20,  # Project
            'N': 30   # Notes
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Format header row (row 1)
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Freeze header row
        worksheet.freeze_panes = 'A2'

        # Format example rows (rows 2-4) with light blue background
        example_fill = PatternFill(start_color='E7E6F7', end_color='E7E6F7', fill_type='solid')
        for row in worksheet['2:4']:
            for cell in row:
                cell.fill = example_fill

        # Format amount columns as currency
        for row in worksheet.iter_rows(min_row=2, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = '$#,##0.00'

        # Center align certain columns
        for col in ['A', 'B', 'D', 'G', 'K']:
            for cell in worksheet[col]:
                if cell.row > 1:
                    cell.alignment = Alignment(horizontal='center', vertical='top')

        # Add borders to all cells
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=len(df)+1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.border = thin_border

        # Create Instructions sheet
        instructions_data = {
            'Section': [
                'INCOME CATEGORIES',
                '', '', '', '', '', '',
                'EXPENSE CATEGORIES',
                '', '', '', '', '', '', '', '',
                'TRANSACTION STATUS',
                '', '', '', '',
                'PAYMENT METHODS',
                '', '', '', '', '',
                'TAX TIPS',
                '', '', '', ''
            ],
            'Details': [
                'Website Package Sale (Essential, Professional, Premium)',
                'Care Plan Subscription (Basic, Professional, Premium)',
                'Down Payment',
                'Final Payment',
                'Add-on Service',
                'Consultation Fee',
                '',
                'Software Subscription (hosting, domains, tools)',
                'Payment Processing Fees',
                'Marketing & Advertising',
                'Office Supplies',
                'Education & Training',
                'Professional Services (legal, accounting)',
                'Equipment',
                'Travel & Meals (business related - 50% deductible)',
                '',
                'Pending - Transaction not yet complete',
                'Completed - Finished and confirmed',
                'Failed - Transaction did not complete',
                'Refunded - Money returned',
                'Stripe',
                'Venmo',
                'Zelle',
                'PayPal',
                'Check',
                'Cash',
                '',
                'Set aside 25-30% of net income for taxes',
                'Pay quarterly estimated taxes if you owe >$1,000',
                'Keep all receipts for 3-7 years',
                'Download Stripe statements monthly',
                'Track business mileage if meeting clients'
            ]
        }

        df_instructions = pd.DataFrame(instructions_data)
        df_instructions.to_excel(writer, sheet_name='Categories & Help', index=False)

        # Format instructions sheet
        instructions_sheet = writer.sheets['Categories & Help']
        instructions_sheet.column_dimensions['A'].width = 25
        instructions_sheet.column_dimensions['B'].width = 60

        # Format header
        for cell in instructions_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Create Summary sheet
        summary_sheet = workbook.create_sheet('Summary')

        summary_sheet['A1'] = 'Hansen Web Services - Financial Summary'
        summary_sheet['A1'].font = Font(bold=True, size=14)

        summary_sheet['A3'] = 'Total Income:'
        summary_sheet['B3'] = '=SUMIF(Transactions!D:D,"Income",Transactions!H:H)'
        summary_sheet['B3'].number_format = '$#,##0.00'
        summary_sheet['B3'].font = Font(bold=True, size=12, color='008000')

        summary_sheet['A4'] = 'Total Expenses:'
        summary_sheet['B4'] = '=SUMIF(Transactions!D:D,"Expense",Transactions!H:H)'
        summary_sheet['B4'].number_format = '$#,##0.00'
        summary_sheet['B4'].font = Font(bold=True, size=12, color='FF0000')

        summary_sheet['A5'] = 'Total Fees Paid:'
        summary_sheet['B5'] = '=SUM(Transactions!I:I)'
        summary_sheet['B5'].number_format = '$#,##0.00'

        summary_sheet['A7'] = 'Net Profit:'
        summary_sheet['B7'] = '=B3-B4'
        summary_sheet['B7'].number_format = '$#,##0.00'
        summary_sheet['B7'].font = Font(bold=True, size=14)
        summary_sheet['B7'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        summary_sheet['A9'] = 'Estimated Tax (30%):'
        summary_sheet['B9'] = '=B7*0.30'
        summary_sheet['B9'].number_format = '$#,##0.00'
        summary_sheet['B9'].font = Font(bold=True, color='FF0000')

        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 20

    print(f"✅ Transaction tracker created successfully!")
    print(f"📁 Location: {output_path}")
    print(f"\nThe Excel file includes:")
    print("  - Transactions sheet (with 3 example entries)")
    print("  - Categories & Help sheet (income/expense categories)")
    print("  - Summary sheet (automatic calculations)")
    print("\nOpen the file in Excel to start tracking your business finances!")

if __name__ == '__main__':
    try:
        create_transaction_tracker()
    except ImportError as e:
        print("❌ Error: Required Python packages not installed")
        print("\nPlease install the required packages:")
        print("  pip install pandas openpyxl")
        print("\nThen run this script again:")
        print("  python3 scripts/create_transaction_tracker.py")
    except Exception as e:
        print(f"❌ Error creating tracker: {e}")
